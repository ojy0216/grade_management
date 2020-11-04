from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
import os
from os import path
from os import system
import math
import copy
import msvcrt

title = "Grade Management"
system("title " + title)
system("mode con cols=115 lines=45")

root = Tk()
root.withdraw()

doc_path = path.expanduser("~\\Documents")
download_path = path.expanduser("~\\Downloads\\noname.xlsx")
log_path = path.join(doc_path, 'GM_dir.dat')  # log 파일 경로


# 백분율성적
def percentage(gpa):
    gpa = math.floor(gpa * 100) / 100  # x.xx 형식으로 내림
    return round(60 + ((gpa - 1) * 40 / 3.5), 2)


def ask_dir():
    while True:
        file_path = filedialog.askopenfilename(
            parent=root, initialdir="/", title='Please select a file')
        if not file_path:  # 취소 선택시
            return -1
        _, extension = path.splitext(file_path)
        if extension == '.xlsx':
            return file_path
        else:
            messagebox.showerror('Error', 'xlsx 파일이 아닙니다.')


def main(original_file=True):
    file_path = None
    need_dir_input = True
    need_log_creation = False
    from_log = False
    from_download = False
    need_manual = True

    while True:
        if path.isfile(download_path):  # download 폴더에 파일이 있는 경우
            flag = messagebox.askquestion('File Found', '원본 파일을 찾았습니다.\n원본 파일을 사용하시겠습니까?')
            if flag == 'yes':
                file_path = download_path
                from_download = True
            else:
                file_path = ask_dir()
                if file_path == -1:
                    return -1
        else:
            if path.isfile(log_path):  # 기존의 파일 위치를 아는 경우
                flag = messagebox.askquestion('File Found', '원본 파일을 찾았습니다.\n원본 파일을 사용하시겠습니까?')
                if flag == 'yes':
                    from_log = True
                    f = open(log_path, 'r')  # 로그 파일에서 파일 경로 읽음
                    file_path = f.readline()
                    _, extension = path.splitext(file_path)
                    if extension == '.xlsx':
                        need_dir_input = False
                    else:
                        os.remove(log_path)  # 로그 파일이 문제있을 경우 파일 삭제
                    f.close()
                else:
                    file_path = ask_dir()
                    if file_path == -1:
                        return -1
                    need_dir_input = False
            if need_dir_input:  # 파일 위치를 모르는 경우 파일 위치를 입력받음
                file_path = ask_dir()
                need_log_creation = True

        if not original_file:
            need_log_creation = False

        wb = load_workbook(file_path)
        ws = wb.active

        file_title = ws['D2'].value
        if (file_title == '성 적 내 역') and need_log_creation:
            f = open(log_path, 'w')  # 파일 위치 저장하는 log 파일 생성
            f.write(file_path)
            f.close()
        elif file_title != '성 적 내 역':  # xlsx 파일은 맞으나 성적파일이 아닐 경우
            if from_log:  # 로그 파일위치의 파일이 문제가 있는 경우
                messagebox.showerror('Error', '{} 파일이 손상된 것 같습니다.\n파일을 다시 다운로드 해주세요.'.format(file_path))
                os.remove(log_path)
                return -1
            if from_download:  # 다운로드 폴더 파일이 문제가 있는 경우
                messagebox.showerror('Error', '{} 파일이 손상된 것 같습니다.\n파일을 다시 다운로드 해주세요.'.format(download_path))
                return -1
            else:
                messagebox.showerror('Error', '성적 파일이 아닙니다.\n유효한 파일을 선택해주세요.')
            continue

        break

    system('cls')
    print()
    print("{0:=^115}".format('[Grade Management]'))
    print()

    xlsx_readout = [row for row in ws.values]  # 열별로 데이터 읽기
    tmp_year = []
    grade_data = []
    my_grade = []
    semester_grade = {
        'credit': 0,  # 학점
        'grade': 0,  # 평점
        'm_credit': 0,  # 전공 학점
        'm_grade': 0,  # 전공 평점
        'wm_credit': 0,  # 전공 학점(전공기초 포함)
        'wm_grade': 0,  # 전공 평점(전공기초 포함)
        'la_credit': 0,  # 교양 학점
        'pf_credit': 0,  # PF 학점
    }
    total_grade = semester_grade
    sem_num = 0

    # 실제 데이터 영역만 grade_data 에 저장
    for row_readout in xlsx_readout:
        if row_readout[0] is not None and row_readout[0] != '년도':
            grade_data.append(row_readout)
            tmp_year.append(row_readout[0])

    year_set = set(tmp_year)  # 수강년도 중 중복 제외
    year_list = list(year_set)
    year_list.sort()  # 수강년도 오름차순 정렬
    sem_list = ['1', '여름', '2', '겨울']

    # 매 학기 별 취득학점, 평균평점, 전공평점, P/F 과목 임시 저장
    for year in year_list:
        for sem in sem_list:
            for grade_row_readout in grade_data:
                if grade_row_readout[0] == year and grade_row_readout[1] == sem:
                    if grade_row_readout[8] != 'P/F':
                        # 학점
                        semester_grade['credit'] += float(grade_row_readout[6])
                        # 평점
                        semester_grade['grade'] += float(grade_row_readout[6]) * float(grade_row_readout[8])
                    if grade_row_readout[2][0:2] == '전공' and not grade_row_readout[2] == '전공기초':
                        # 전공 학점
                        semester_grade['m_credit'] += float(grade_row_readout[6])
                        # 전공 평점
                        semester_grade['m_grade'] += float(grade_row_readout[6]) * float(grade_row_readout[8])
                    if grade_row_readout[2][0:2] == '전공':
                        # 전공 학점(전공기초 포함)
                        semester_grade['wm_credit'] += float(grade_row_readout[6])
                        # 전공 평점(전공기초 포함)
                        semester_grade['wm_grade'] += float(grade_row_readout[6]) * float(grade_row_readout[8])
                    if grade_row_readout[2] == '교양':
                        semester_grade['la_credit'] += float(grade_row_readout[6])
                    if grade_row_readout[8] == 'P/F':
                        # PF 학점
                        semester_grade['pf_credit'] += float(grade_row_readout[6])
            my_grade.append(copy.deepcopy(semester_grade))

            # 학기 별 취득학점, 평균평점, 전공평점 출력
            if semester_grade['credit'] + semester_grade['pf_credit'] != 0:
                if sem == '1' or sem == '2':
                    sem_num += 1
                    sem_string = '({}차 학기)'.format(sem_num)
                else:
                    sem_string = ''.rjust(7)
                print('-' * 115)
                print(" {}년 {}학기 ".format(year, sem) + sem_string, end='  ||  ')
                if semester_grade['credit'] == 0:
                    grade = 'N/A'
                else:
                    grade = round(semester_grade['grade'] / semester_grade['credit'], 3)
                if semester_grade['m_credit'] == 0:
                    major_grade = 'N/A'
                else:
                    major_grade = round(semester_grade['m_grade'] / semester_grade['m_credit'], 3)
                if semester_grade['wm_credit'] == 0:
                    wide_major_grade = 'N/A'
                else:
                    wide_major_grade = round(semester_grade['wm_grade'] / semester_grade['wm_credit'], 3)
                print("취득학점: {} | 평균평점: {} | 전공평점: {} | 전공평점(전공기초 포함): {}".
                      format(str(semester_grade['credit'] + semester_grade['pf_credit']).ljust(4), str(grade).ljust(5),
                             str(major_grade).ljust(5), str(wide_major_grade).ljust(4)))

            # 매 학기 별 초기화
            for key in semester_grade.keys():
                semester_grade[key] = 0

    # 총 성적에 학기 별 성적 누적
    for my_grade_row_readout in my_grade:
        for key in my_grade_row_readout.keys():
            total_grade[key] += my_grade_row_readout[key]

    total_credit = total_grade['credit'] + total_grade['pf_credit']
    ave_grade = round(total_grade['grade'] / total_grade['credit'], 3)
    try:
        ave_major_grade = round(total_grade['m_grade'] / total_grade['m_credit'], 3)
    except ZeroDivisionError:
        ave_major_grade = 'N/A'
    try:
        ave_wide_major_grade = round(total_grade['wm_grade'] / total_grade['wm_credit'], 3)
    except ZeroDivisionError:
        ave_wide_major_grade = 'N/A'

    print('=' * 115)
    print("\t총취득학점: {} | 총취득전공학점: {} | 총취득전공학점(전공기초 포함): {} | 총취득교양학점: {}".format(
        str(total_credit).center(5),
        str(total_grade['m_credit']).ljust(5),
        str(total_grade['wm_credit']).ljust(5),
        str(total_grade['la_credit']).ljust(5)
    ))
    print('-' * 115)
    print("\t총평균평점: {} | 총평균전공평점: {} | 총평균전공평점(전공기초 포함): {} | GPA: {}".format(
        str(ave_grade).ljust(5),
        str(ave_major_grade).ljust(5),
        str(ave_wide_major_grade).ljust(5),
        percentage(ave_grade)
    ))
    print('=' * 115)

    sim_total_grade = copy.deepcopy(total_grade)

    # 미래 성적 계산
    while True:
        flag = input("\n 다음 학기 Simulation(Y/N, M for manual, R for reset): ").upper()

        if flag not in ('Y', 'M', 'R'):
            break

        if flag == 'R':  # Reset
            sim_total_grade = copy.deepcopy(total_grade)
            print('-' * 115)
            print(" [현재 시점]  총취득학점 : {} | 총평균평점: {} | 총평균전공평점: {} | GPA: {}".
                  format(total_credit, ave_grade, ave_major_grade, percentage(ave_grade)))
            print('-' * 115)
            continue

        elif flag == 'M':  # Manual
            sim_credit = 0
            sim_grade = 0
            tmp_sim_grade = 0
            grade_dict = {
                'A+': 4.5,
                'A': 4,
                'B+': 3.5,
                'B': 3,
                'C+': 2.5,
                'C': 2,
                'D+': 1.5,
                'D': 1,
                'F': 0,
                'P': 0
            }
            if need_manual:
                messagebox.showinfo('Manual Input',
                                    '학점과 성적을 공백으로 구분하여\n하나씩 입력해주세요.\n\n입력 예시:\n[1]: 3 a+\n[2]: 2 B\n\n-1 을 입력하여 종료')
                need_manual = False
            i = 1
            while True:
                tmp = input(' [{}]: '.format(i))
                try:
                    tmp_credit, tmp_grade = tmp.split()
                    tmp_grade = tmp_grade.upper()
                except ValueError:
                    try:
                        if float(tmp) <= 0:
                            sim_grade = tmp_sim_grade / sim_credit
                            print('-' * 115)
                            print(' [입력] 학점: {} | 평균평점: {}'.format(sim_credit, sim_grade))
                            break
                    except ValueError:
                        messagebox.showwarning('Warning', '학점 입력 오류')
                        continue
                    messagebox.showwarning('Warning', '입력 오류')
                    continue

                try:
                    tmp_credit = float(tmp_credit)
                except ValueError:
                    messagebox.showwarning('Warning', '학점 입력 오류')
                    continue

                if tmp_grade in grade_dict.keys():
                    if tmp_grade == 'P':
                        sim_total_grade['pf_credit'] += sim_credit
                    else:
                        sim_credit += tmp_credit
                        tmp_sim_grade += grade_dict[tmp_grade] * tmp_credit
                    i += 1
                else:
                    messagebox.showwarning('Warning', '평점 입력 오류')
                    continue

        else:  # Only credit, grade
            try:
                sim_credit = float(input(" 희망 학점: "))
                if sim_credit <= 0:
                    messagebox.showwarning('Warning', '학점 입력 오류')
                    continue
            except ValueError:
                messagebox.showwarning('Warning', '학점 입력 오류')
                continue

            try:
                sim_grade = float(input(" 희망 평점: "))
                if (sim_grade < 0) or (sim_grade > 4.5):
                    messagebox.showwarning('Warning', '평점 입력 오류: 입력 가능 범위 = 0 ~ 4.5')
                    continue
            except ValueError:
                messagebox.showwarning('Warning', '평점 입력 오류')
                continue

        previous_ave_grade = round(sim_total_grade['grade'] / sim_total_grade['credit'], 3)
        sim_total_grade['credit'] += sim_credit
        sim_total_grade['grade'] += sim_grade * sim_credit

        new_total_credit = sim_total_grade['credit'] + sim_total_grade['pf_credit']
        new_ave_grade = round(sim_total_grade['grade'] / sim_total_grade['credit'], 3)

        print('-' * 115)
        print(" 총취득학점: {0}({1:+.01f}) | 총평균평점: {2}({3:+.03f}) | GPA: {4}({5:+.02f})".
              format(new_total_credit, sim_credit,
                     new_ave_grade, round(new_ave_grade - previous_ave_grade, 3),
                     percentage(new_ave_grade), round(percentage(new_ave_grade) - percentage(previous_ave_grade), 2)))
        print('-' * 115)

    return file_path


if __name__ == '__main__':
    file_path = main()
    while True:
        if file_path == -1:
            break
        modify_file = input('\n Want to modify grade? (Y/N): ').upper()
        if modify_file == 'Y':
            os.startfile(file_path)
        restart = input('\n Want to restart program? (Y/N): ').upper()
        if restart == 'Y':
            user_exit = main(original_file=False)
            if user_exit == -1:
                break
        else:
            break

    print("\n Press any key to exit")
    msvcrt.getch()
